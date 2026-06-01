"""표창현황 액셀 → DB import.

사용:
    cd backend
    source .venv/bin/activate
    python scripts/import_status_excel.py

규칙:
- 박중보 시트: 전체 row 모두 처리 (모든 의원)
- 홍서희·이민우·김수경(휴직) 시트: 민주 의원의 row만 처리 (사용자 요청)
- title prefix '[액셀]'로 표시 → 재실행 시 같은 prefix 가진 case 모두 삭제 후 새로 import
- 날짜 'M.D.' 형식: 7~12월 → 2025, 1~6월 → 2026
- 'YY.M.D.' 형식: 그대로
- 표창대상자: 콤마/줄바꿈으로 split → 각 Recipient 생성
"""
from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

# 프로젝트 루트 추가
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.database import SessionLocal, init_db
from app.legislators import LEGISLATORS
from app.models import AwardCase, Recipient

EXCEL_PATH = (
    "/Users/jun/Desktop/클로드 코드/업무 개선 프로젝트/표창/"
    "★보건복지전문위원회 표창현황 ver (홍서희).xlsx"
)
TITLE_PREFIX = "[액셀] "

VALID_STATUSES = {"대기", "예정", "진행", "보관", "완료", "취소"}

DEMOCRATIC_NAMES = {L.name for L in LEGISLATORS if L.party == "민주"}

# 시트별 컬럼 매핑 (row 인덱스 — A열은 비어있어 row[0]=None)
# 박중보·홍서희·이민우 동일 구조; 김수경(휴직) 시트는 컬럼이 2개 적음
SHEET_CONFIGS = {
    "박중보": {
        "filter": "all",
        "seq": 1, "status": 2, "recommender": 5, "organization": 6,
        "submit_date": 8, "award_date": 11, "contact_name": 12,
        "contact_phone": 14, "address": 16, "recipient": 17,
    },
    "홍서희": {
        "filter": "democratic",
        "seq": 1, "status": 2, "recommender": 5, "organization": 6,
        "submit_date": 8, "award_date": 11, "contact_name": 12,
        "contact_phone": 14, "address": 16, "recipient": 17,
    },
    "이민우": {
        "filter": "democratic",
        "seq": 1, "status": 2, "recommender": 5, "organization": 6,
        "submit_date": 8, "award_date": 11, "contact_name": 12,
        "contact_phone": 14, "address": 16, "recipient": 17,
    },
    "김수경(휴직)": {
        "filter": "democratic",
        # 조서작성·상신확인 컬럼이 없어 의원명이 D열로 앞당겨짐
        "seq": 1, "status": 2, "recommender": 3, "organization": 4,
        "submit_date": 6, "award_date": 9, "contact_name": 10,
        "contact_phone": 12, "address": 14, "recipient": 15,
    },
}


def parse_md_date(s):
    """문자열 날짜를 date로. 'M.D.' 형식이면 7~12=2025, 1~6=2026. 'YY.M.D.'는 그대로."""
    if s is None:
        return None
    if isinstance(s, date):
        return s
    s = str(s).strip().rstrip(".")
    if not s:
        return None
    parts = s.split(".")
    parts = [p.strip() for p in parts if p.strip()]
    try:
        if len(parts) == 2:
            m, d = int(parts[0]), int(parts[1])
            year = 2025 if m >= 7 else 2026
            return date(year, m, d)
        if len(parts) == 3:
            yy, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            year = 2000 + yy if yy < 100 else yy
            return date(year, m, d)
    except (ValueError, TypeError):
        return None
    return None


def parse_recipients(s):
    """표창대상자 셀(R열)을 콤마/줄바꿈으로 split. _x000D_(엑셀 줄바꿈) 처리."""
    if not s:
        return []
    s = str(s).replace("_x000D_", "\n")
    parts = re.split(r"[,，\n]", s)
    return [p.strip() for p in parts if p.strip() and p.strip() != "등"]


def normalize_status(s):
    if not s:
        return "예정"
    s = str(s).strip()
    return s if s in VALID_STATUSES else "예정"


def process_sheet(db, ws, cfg, sheet_name) -> tuple[int, int]:
    case_count = 0
    recipient_count = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not any(row):
            continue

        def get(key):
            idx = cfg[key]
            return row[idx] if idx < len(row) else None

        recommender = get("recommender")
        if not recommender or not isinstance(recommender, str):
            continue
        recommender = recommender.strip()
        if not recommender:
            continue

        # 시트별 필터 — democratic이면 민주 의원만
        if cfg["filter"] == "democratic" and recommender not in DEMOCRATIC_NAMES:
            continue

        organization = get("organization")
        recipient_str = get("recipient")
        recipients_names = parse_recipients(recipient_str)
        if not recipients_names:
            if organization and isinstance(organization, str):
                recipients_names = [organization.strip()]
            else:
                continue

        rec_date = parse_md_date(get("submit_date"))
        aw_date = parse_md_date(get("award_date"))
        status = normalize_status(get("status"))
        contact_name = get("contact_name")
        contact_phone = get("contact_phone")
        address = get("address")

        title = f"{TITLE_PREFIX}{recommender} 의원 — {organization or ''}".strip()
        case = AwardCase(
            title=title[:255],
            award_grade="경기도의회 의장 표창",
            recommender_department="보건복지위원회",
            recommender_position="위원",
            recommender_name=recommender,
            recommender_full_title=f"경기도의회 보건복지위원회 의원   {recommender}",
            recommendation_date=rec_date,
            award_date=aw_date,
            applicant_name=(str(contact_name).strip() if contact_name else None),
            applicant_contact=(str(contact_phone).strip() if contact_phone else None),
            applicant_delivery_address=(str(address).strip() if address else None),
            status=status,
        )
        db.add(case)
        db.flush()

        for idx, name in enumerate(recipients_names, start=1):
            r = Recipient(
                award_case_id=case.id,
                sequence_no=idx,
                recipient_name=name[:255],
                organization_name=(
                    str(organization).strip() if organization else None
                ),
                recommendation_rank="1순위",
            )
            db.add(r)
            recipient_count += 1
        case_count += 1
    return case_count, recipient_count


def main():
    init_db()
    db = SessionLocal()
    try:
        # 설정 초기화 후 의원 명단이 비어 있으면 기본 명단(보건복지 의원)을 시드
        # — 쿼터 현황에 의원 행이 보이도록. (실험/재import 편의)
        from app.models import Legislator
        if db.query(Legislator).count() == 0:
            for i, L in enumerate(LEGISLATORS):
                db.add(
                    Legislator(
                        name=L.name, party=L.party, is_chair=L.is_chair,
                        staff=L.staff, seal_filename=L.seal_filename,
                        sort_order=i, active=True,
                    )
                )
            db.commit()
            print(f"의원 명단 {len(LEGISLATORS)}명 시드")

        # 기존 [액셀] prefix case 삭제
        existing = (
            db.query(AwardCase).filter(AwardCase.title.like(f"{TITLE_PREFIX}%")).all()
        )
        for c in existing:
            db.delete(c)
        if existing:
            print(f"기존 액셀 import case {len(existing)}건 삭제")
        db.commit()

        wb = load_workbook(EXCEL_PATH, data_only=True)

        total_case = 0
        total_recipient = 0
        sheet_summary = []

        for sheet_name, cfg in SHEET_CONFIGS.items():
            if sheet_name not in wb.sheetnames:
                print(f"시트 '{sheet_name}'가 없습니다. skip")
                continue
            ws = wb[sheet_name]
            case_count, recipient_count = process_sheet(db, ws, cfg, sheet_name)
            total_case += case_count
            total_recipient += recipient_count
            sheet_summary.append((sheet_name, case_count, recipient_count, cfg["filter"]))

        db.commit()
        print("\n=== import 결과 ===")
        for sn, cc, rc, f in sheet_summary:
            tag = "전체" if f == "all" else "민주 의원만"
            print(f"  {sn:<14} ({tag:<10}) case {cc:>4}건, recipient {rc:>4}명")
        print(f"  {'합계':<14} {'':>12} case {total_case:>4}건, recipient {total_recipient:>4}명")
    finally:
        db.close()


if __name__ == "__main__":
    main()
