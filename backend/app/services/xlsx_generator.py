"""XLSX 생성 서비스
- 01. 공적개요서.xlsx
- 03. 표창대상자.xlsx
"""
from __future__ import annotations

import re
from copy import copy as _copy_style
from datetime import date
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import GENERATED_DIR
from ..models import AwardCase, Recipient


def _extract_region_from_address(address: str | None) -> str:
    """주소에서 'OO시' 추출 — 예: '경기도 수원시 영통구 ...' -> '수원시'.
    '서울특별시'/'부산광역시' 등은 '서울'/'부산'으로 축약."""
    if not address:
        return ""
    m = re.search(r"[가-힣]+시", address)
    if not m:
        return ""
    region = m.group()
    for suffix in ("특별시", "광역시", "특별자치시"):
        if suffix in region:
            return region.replace(suffix, "")
    return region


def _yymmdd_from_birth_date(d) -> str:
    if not d or not isinstance(d, date):
        return ""
    return f"{d.year % 100:02d}{d.month:02d}{d.day:02d}"

_THIN = Side(border_style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_HDR_FILL = PatternFill("solid", fgColor="F3F4F6")


def _fmt_date_dot(d) -> str:
    if not d:
        return ""
    if isinstance(d, date):
        return f"{d.year:04d}.{d.month:02d}.{d.day:02d}."
    return str(d)


def _merit_overview_text(recipient: Recipient) -> str:
    mc = recipient.merit_content
    if not mc:
        return ""
    lines = []
    for idx, key in enumerate(
        ["merit_overview_1", "merit_overview_2", "merit_overview_3", "merit_overview_4"],
        start=1,
    ):
        val = getattr(mc, key) or ""
        if val.strip():
            lines.append(f"{idx}. {val.strip()}")
    return "\n".join(lines)


def generate_merit_overview_xlsx(case: AwardCase) -> Path:
    """01. 공적개요서.xlsx 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "공적개요서"

    # 제목
    ws.merge_cells("A1:G1")
    ws["A1"] = "공 적 개 요 서"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 36

    headers = ["연번", "단체명", "성 명\n(생년월일)", "공적분야", "추천훈격", "직위", "공 적 개 요"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.fill = _HDR_FILL
        cell.border = _BORDER
    ws.row_dimensions[3].height = 36

    # 컬럼 폭
    widths = [6, 14, 18, 14, 14, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    for r in case.recipients or []:
        name_dob = f"{r.recipient_name}\n({_fmt_date_dot(r.birth_date)})"
        values = [
            r.sequence_no or row - 3,
            r.organization_name or "",
            name_dob,
            r.merit_category or "",
            case.award_grade or "",
            r.recipient_position_title or "",
            _merit_overview_text(r),
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = _CENTER if col != 7 else _LEFT
            cell.border = _BORDER
        ws.row_dimensions[row].height = 110
        row += 1

    suffix = case.recommender_name or "추천자"
    file_name = f"01. 공적개요서({suffix} 의원).xlsx"
    file_path = GENERATED_DIR / file_name
    wb.save(file_path)
    return file_path


# 사용자 정답 샘플(03 표창대상자)을 그대로 템플릿으로 사용 → 색·폰트·테두리·병합·
# 너비 등 모든 서식을 100% 보존한다. 데이터(PII)는 비워둔 상태로 저장돼 있다.
_RECIPIENT_TEMPLATE = (
    Path(__file__).resolve().parent.parent / "templates" / "표창대상자_template.xlsx"
)


def _clone_row_style(ws, src_row: int, dst_row: int) -> None:
    """src_row 각 셀의 서식(폰트·채움·테두리·정렬·표시형식·행높이)을 dst_row로 복제."""
    for col in range(1, 15):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        d.font = _copy_style(s.font)
        d.fill = _copy_style(s.fill)
        d.border = _copy_style(s.border)
        d.alignment = _copy_style(s.alignment)
        d.number_format = s.number_format
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _recipient_list_title(year: int, award_grade: str | None) -> str:
    return f"{year}년도 {award_grade or '표창'} 추천자 명단"


def generate_recipient_list_template_xlsx(
    award_grade: str = "경기도의회 의장 표창", empty_rows: int = 15
) -> Path:
    """빈 표창대상자 업로드 서식 — 샘플 템플릿 서식 그대로, 빈 입력 행 N개."""
    wb = load_workbook(_RECIPIENT_TEMPLATE)
    ws = wb.active
    ws["A1"] = _recipient_list_title(date.today().year, award_grade)
    # 행4(스타일 보존된 빈 행) 기준으로 빈 입력 행 확장
    for row in range(5, 4 + max(1, empty_rows)):
        _clone_row_style(ws, 4, row)
    file_path = GENERATED_DIR / "표창대상자_업로드서식.xlsx"
    wb.save(file_path)
    return file_path


def generate_recipient_list_xlsx(case: AwardCase) -> Path:
    """03. 표창대상자.xlsx — 샘플 템플릿 서식 그대로 데이터만 채움."""
    wb = load_workbook(_RECIPIENT_TEMPLATE)
    ws = wb.active
    year = case.award_date.year if case.award_date else date.today().year
    ws["A1"] = _recipient_list_title(year, case.award_grade)

    recipients: List[Recipient] = list(case.recipients or [])
    for i, r in enumerate(recipients):
        row = 4 + i
        if row > 4:  # 둘째 대상자부터: 행4 서식을 복제
            _clone_row_style(ws, 4, row)
        award_dt = case.award_date  # date 객체 (템플릿 M열 표시형식 그대로)
        recommender_dept = case.recommender_department or ""
        recommender_pos = case.recommender_position or "위원"
        region = r.region or _extract_region_from_address(r.address)
        birth_yymmdd = r.birth_yymmdd or _yymmdd_from_birth_date(r.birth_date)
        values = [
            r.sequence_no or i + 1,
            recommender_dept,
            recommender_pos,
            case.recommender_name or "",
            region,
            r.address or "",
            r.organization_name or "",
            r.recipient_position_title or "",
            r.recipient_name or "",
            birth_yymmdd,
            r.merit_category or "",
            r.merit_period or "",
            award_dt if award_dt else "",
            r.note or "",
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)

    # 파일명 규칙: 03. 표창대상자(000 의원 추천_홍길동 등 N인).xlsx
    recommender = case.recommender_name or "추천자"
    recipients: List[Recipient] = list(case.recipients or [])
    first_name = recipients[0].recipient_name if recipients else "대상자"
    n = len(recipients) or 1
    file_name = f"03. 표창대상자({recommender} 의원 추천_{first_name} 등 {n}인).xlsx"
    file_path = GENERATED_DIR / file_name
    wb.save(file_path)
    return file_path
