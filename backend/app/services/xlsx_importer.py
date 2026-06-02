"""기존 표창대상자 XLSX 업로드 → DB 반영 (3차 기능)"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from ..models import AwardCase, Recipient


def _to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, date):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    return str(v).strip()


def _parse_birth(yymmdd: str) -> tuple[date | None, str]:
    s = "".join(ch for ch in yymmdd if ch.isdigit())[:6]
    if len(s) != 6:
        return None, s
    yy, mm, dd = int(s[:2]), int(s[2:4]), int(s[4:6])
    year = 2000 + yy if yy < 30 else 1900 + yy
    try:
        return date(year, mm, dd), s
    except ValueError:
        return None, s


def import_recipients_from_xlsx(file_path: Path, case: AwardCase) -> List[Recipient]:
    """업로드된 표창대상자 양식(03번 양식) → Recipient 리스트로 변환"""
    wb = load_workbook(filename=str(file_path), data_only=True)
    ws = wb.active

    recipients: List[Recipient] = []
    # 헤더가 1~3행에 걸쳐 있으므로 4행부터 데이터
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or all(c is None for c in row):
            continue
        # 컬럼: A연번 B추천소속 C추천직위 D추천성명 E지역 F주소
        # G대상소속 H대상직위/직명 I대상성명 J생년월일6 K공적분야 L공적기간 M표창일 N비고
        (seq, rdept, rpos, rname, region, addr, org, pos_title, name,
         birth6, field, period, award_dt, note) = (list(row) + [None] * 14)[:14]
        if not name:
            continue

        birth_date, birth_yy = _parse_birth(_to_str(birth6))

        # 표창일(M열) — 대상자 개인 단위로 저장. datetime이면 date로 변환.
        if isinstance(award_dt, datetime):
            r_award_date = award_dt.date()
        elif isinstance(award_dt, date):
            r_award_date = award_dt
        else:
            r_award_date = None

        r = Recipient(
            sequence_no=int(seq) if seq else len(recipients) + 1,
            recipient_name=_to_str(name),
            birth_date=birth_date,
            birth_yymmdd=birth_yy,
            address=_to_str(addr),
            region=_to_str(region),
            organization_name=_to_str(org),
            recipient_position_title=_to_str(pos_title),
            merit_category=_to_str(field),
            merit_period=_to_str(period),
            award_date=r_award_date,
            note=_to_str(note),
        )
        recipients.append(r)

        # 추천자 정보가 첫 행에 있으면 case에 반영
        if rname and not case.recommender_name:
            case.recommender_department = _to_str(rdept)
            case.recommender_position = _to_str(rpos)
            case.recommender_name = _to_str(rname)
        # case.award_date는 대표/폴백값(첫 행 날짜) — 목록 정렬·미설정 대상자 폴백용
        if r_award_date and not case.award_date:
            case.award_date = r_award_date

    return recipients
